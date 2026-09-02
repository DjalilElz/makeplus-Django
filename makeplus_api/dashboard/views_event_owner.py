"""
Registration submissions view for event owners.

Accessible by staff/superusers (any event) and by event-owner accounts
(events.UserEventAssignment role='event_owner') scoped to the event(s)
they're assigned to. Owners can view, search and filter (by registration
status, by bloc item, or by free text), edit a free-text note per
registration, move a registration between its 6 statuses (Registered /
Reserved / To Contact / To Recontact / Confirmed / Cancelled -- see
caisse.services for what Confirmed/Cancelled actually do), and delete a
registration.
"""
import json
from collections import Counter
from decimal import Decimal

from types import SimpleNamespace

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.exceptions import PermissionDenied
from django.db.models import Q
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.views.decorators.cache import never_cache
from django.views.decorators.http import require_POST

from caisse.services import (
    cancel_registration_order, confirm_registration_order, resync_confirmed_registration_order,
    unconfirm_registration_order,
)
from dashboard.email_sender import send_email
from dashboard.models_email import get_event_email_template
from events.models import Event, ParticipantEventRegistration, UserEventAssignment
from .blocs_service import (
    compute_order, serialize_status_rules_for_period, serialize_period_baseline_rules,
)
from .models_blocs import RegistrationOrder
from .views_blocs import get_public_bloc_context

BLOC_KEYS = ('status', 'restauration', 'workshops', 'social_event')


def _empty_bloc_context():
    """Safe fallback for info-only events or events with no visible blocs."""
    return {
        'has_blocs': False,
        'config': None,
        'custom_blocs': [],
        'custom_blocs_before_workshops': [],
        'custom_blocs_after_workshops': [],
        'workshops_visible': False,
        'paid_sessions': [],
        'active_bloc_keys': [],
        'status_rules_json': '{}',
        'period_rules_json': '{}',
        'periods': [],
        'current_period_id': '',
        'all_periods_rules_json': '{}',
    }


def _redirect_to_submissions(request, event_id):
    """
    Back to the submissions list, carrying over whatever status/search/
    item filters were active (request.GET) -- a plain redirect() drops
    them, silently resetting the owner's filtered view to "everything"
    after every status change/delete, on top of losing their scroll
    position (which the page's own JS restores separately).
    """
    url = reverse('dashboard:event_owner_submissions', kwargs={'event_id': event_id})
    query = request.GET.urlencode()
    if query:
        url = f'{url}?{query}'
    return redirect(url)


def _can_access_event_orders(user, event):
    """Staff/superusers can access any event; event owners only their own."""
    if user.is_staff or user.is_superuser:
        return True
    return UserEventAssignment.objects.filter(
        user=user, event=event, role='event_owner', is_active=True
    ).exists()


@never_cache
@login_required
def event_owner_submissions_home(request):
    """
    Landing page for event owners after login. Redirects straight to
    their event if they own only one, otherwise shows a picker.
    """
    assignments = UserEventAssignment.objects.filter(
        user=request.user, role='event_owner', is_active=True
    ).select_related('event').order_by('-assigned_at')

    if not assignments.exists():
        if request.user.is_staff or request.user.is_superuser:
            return redirect('dashboard:home')
        messages.error(request, "Vous n'avez accès à aucun événement en tant qu'organisateur.")
        return redirect('dashboard:login')

    if assignments.count() == 1:
        return redirect('dashboard:event_owner_submissions', event_id=assignments.first().event_id)

    events = [a.event for a in assignments]
    return render(request, 'dashboard/event_owner/event_picker.html', {'events': events})


STATUS_LABELS = {
    'pending': 'Enregistré',
    'reserved': 'Réservé',
    'to_contact': 'Contacter',
    'to_recontact': 'À recontacter',
    'approved': 'Confirmé',
    'rejected': 'Annulé',
}


def _bloc_names(order):
    """Item name(s) picked in each bloc for this order, comma-joined
    (a bloc like workshops can have more than one)."""
    names = {b: [] for b in BLOC_KEYS}
    for it in order.items_snapshot:
        b = it.get('bloc')
        if b in names:
            names[b].append(it.get('name', ''))
    return {b: ', '.join(v) if v else '—' for b, v in names.items()}


def _bloc_item_ids(order, bloc):
    return {str(it.get('id')) for it in order.items_snapshot if it.get('bloc') == bloc}


def _blocs_editor_selection_json(order):
    """
    This order's current selection, split into the two shapes the blocs
    editor modal's form fields expect (BlocItem ids vs. Session ids) --
    used to pre-check the right boxes when an owner opens the modal.
    """
    items = [str(it.get('id')) for it in order.items_snapshot if it.get('type') == 'item']
    sessions = [str(it.get('id')) for it in order.items_snapshot if it.get('type') == 'session']
    return json.dumps({'items': items, 'sessions': sessions})


def _period_rules_by_period_json(event, orders):
    """
    Status/period BlocItemStatusRule payloads, one set per distinct
    ReductionPeriod actually used among these orders (plus an entry for
    "no period"), keyed by period id (or '' for none).

    The blocs-editor popup's live price preview must match what
    registration_order_blocs_save() will actually save -- which is now
    pinned to the ORDER's own snapshotted period, not today's -- so the
    popup needs each order's own period's rules, not one shared "today"
    ruleset for the whole page.
    """
    periods = {order.period for order in orders}  # includes None if any order has no period
    payload = {}
    for period in periods:
        key = str(period.id) if period else ''
        payload[key] = {
            'status': serialize_status_rules_for_period(event, period),
            'period': serialize_period_baseline_rules(event, period),
        }
    return json.dumps(payload)


def _free_form_submission_rows(event):
    """Expose free-form event submissions as lightweight rows for info-only events."""
    from .models_form import FormSubmission

    rows = []
    form_config = event.custom_forms.first()
    if not form_config:
        return rows

    phone_field_name = next(
        (f.get('name') for f in form_config.fields_config if f.get('type') == 'tel'),
        None,
    )

    for submission in FormSubmission.objects.filter(form__event=event).order_by('-submitted_at'):
        data = submission.data or {}
        full_name = (
            data.get('full_name')
            or ' '.join(filter(None, [data.get('first_name', ''), data.get('last_name', '')]))
            or submission.email
            or ''
        )
        row = SimpleNamespace(
            id=submission.id,
            event=event,
            period=None,
            period_id=None,
            full_name=full_name,
            email=submission.email or data.get('email') or '',
            phone=(data.get(phone_field_name, '') if phone_field_name else ''),
            items_snapshot=[],
            total_before_reduction=Decimal('0'),
            total_after_reduction=Decimal('0'),
            total_discount_percent=Decimal('0'),
            status='pending',
            status_label='Enregistré',
            admin_notes=submission.admin_notes,
            reviewed_by=None,
            reviewed_by_caisse=None,
            reviewed_at=None,
            payment_link_sent_at=None,
            receipt_file=None,
            form_submission=submission,
            created_at=submission.submitted_at,
            duplicate_count=0,
            is_duplicate_email=False,
            bloc_names={'status': '—', 'restauration': '—', 'workshops': '—', 'social_event': '—'},
            participant_status='—',
            blocs_editor_selection_json='{"items": [], "sessions": []}',
            blocs_editor_period_key='',
            blocs_discount_percent=Decimal('0'),
            period_discount_percent=Decimal('0'),
        )
        rows.append(row)
    return rows


@never_cache
@login_required
def event_owner_submissions(request, event_id):
    """List registration submissions for one event, read-only."""
    event = get_object_or_404(Event, id=event_id)

    if not _can_access_event_orders(request.user, event):
        raise PermissionDenied("You do not have access to this event's submissions.")

    status = request.GET.get('status', '').strip()
    query = request.GET.get('q', '').strip()
    item_filters = {
        'status': request.GET.get('item_status', '').strip(),
        'restauration': request.GET.get('item_restauration', '').strip(),
        'social_event': request.GET.get('item_social_event', '').strip(),
    }
    workshop_item_ids = [v for v in request.GET.getlist('item_workshops') if v]

    # reviewed_by / reviewed_by_caisse are dereferenced per row in the
    # template; without eager-loading them here every order fires two extra
    # queries, and on a large event that N+1 blows past the gunicorn worker
    # timeout (seen in prod as WORKER TIMEOUT -> 500 on this page).
    all_orders = list(
        RegistrationOrder.objects.filter(event=event).select_related(
            'period', 'form_submission__form',
            'reviewed_by', 'reviewed_by_caisse',
        ).order_by('-created_at')
    )
    if not event.bloc_config or not event.bloc_config.any_bloc_visible():
        all_orders.extend(_free_form_submission_rows(event))

    # Filter dropdown options are built from the event's full, unfiltered
    # order set so choosing a filter never makes other options disappear.
    bloc_item_options = {b: {} for b in BLOC_KEYS}
    for order in all_orders:
        for it in order.items_snapshot:
            b = it.get('bloc')
            if b in bloc_item_options:
                bloc_item_options[b][str(it.get('id'))] = it.get('name', '')
    for b in bloc_item_options:
        bloc_item_options[b] = sorted(bloc_item_options[b].items(), key=lambda pair: pair[1] or '')

    # items_snapshot is a JSON field, not a normalized relation -- item
    # filters are applied in Python rather than at the DB query level.
    orders = all_orders
    if status in STATUS_LABELS:
        orders = [o for o in orders if o.status == status]
    if query:
        ql = query.lower()
        orders = [o for o in orders if ql in (o.full_name or '').lower() or ql in (o.email or '').lower()]
    for bloc, item_id in item_filters.items():
        if item_id:
            orders = [o for o in orders if item_id in _bloc_item_ids(o, bloc)]
    if workshop_item_ids:
        # A person can pick several workshops -- "picked any of these",
        # not "picked all of these" (AND would too easily match nobody).
        wanted = set(workshop_item_ids)
        orders = [o for o in orders if _bloc_item_ids(o, 'workshops') & wanted]

    # A phone number, if the event's registration form has a "tel"-type
    # custom field, lives in FormSubmission.data under whatever name the
    # admin gave that field -- not a fixed column. Every order for this
    # event shares the same form, so resolve the field name once.
    phone_field_name = None
    form_config = event.custom_forms.first()
    if form_config:
        phone_field_name = next(
            (f.get('name') for f in form_config.fields_config if f.get('type') == 'tel'), None,
        )

    # Same email submitting more than once for this event -- flagged in the
    # template so the owner can spot and handle duplicates manually.
    # Counted against the event's full, unfiltered order set so a status/
    # item filter doesn't hide a duplicate's sibling from the count.
    email_counts = Counter((o.email or '').strip().lower() for o in all_orders if o.email)

    total_revenue = Decimal('0')
    status_counts = {code: 0 for code in STATUS_LABELS}
    item_counts = {b: {} for b in BLOC_KEYS}
    for order in orders:
        order.bloc_names = _bloc_names(order)
        order.participant_status = order.bloc_names['status']
        order.status_label = STATUS_LABELS.get(order.status, order.status)
        order.phone = (
            order.form_submission.data.get(phone_field_name, '')
            if order.form_submission and phone_field_name else ''
        )
        order.duplicate_count = email_counts.get((order.email or '').strip().lower(), 0)
        order.is_duplicate_email = order.duplicate_count > 1
        order.blocs_editor_selection_json = _blocs_editor_selection_json(order)
        order.blocs_editor_period_key = str(order.period_id) if order.period_id else ''
        total_revenue += order.total_after_reduction
        status_counts[order.status] = status_counts.get(order.status, 0) + 1
        for it in order.items_snapshot:
            b = it.get('bloc')
            if b in item_counts:
                name = it.get('name') or '—'
                item_counts[b][name] = item_counts[b].get(name, 0) + 1

    item_counts = {b: sorted(counts.items(), key=lambda pair: -pair[1]) for b, counts in item_counts.items()}
    status_breakdown = [(code, STATUS_LABELS[code], count) for code, count in status_counts.items()]

    # Registration pace over time, for a trend chart -- one point per day
    # that had at least one submission, in chronological order.
    day_counts = Counter(order.created_at.date().isoformat() for order in orders)
    registrations_by_day = sorted(day_counts.items())

    # "Mes événements" only makes sense when the owner actually has more than
    # one event to switch between; with a single event there's nothing to go
    # back to, so we hide it and just offer logout. Staff keep the link as a
    # way back to their main dashboard.
    owner_events_count = UserEventAssignment.objects.filter(
        user=request.user, role='event_owner', is_active=True
    ).count()
    show_events_link = request.user.is_staff or request.user.is_superuser or owner_events_count > 1

    # Same helper the public registration form itself uses -- the "edit
    # blocs" modal renders this exact catalog/pricing data, so what an
    # owner sees and picks from is byte-identical to what a participant
    # sees, not a separately-maintained lookalike. include_inactive=True
    # is the one deliberate difference: unlike the public form, the owner
    # can still see and toggle items that have since been deactivated
    # (e.g. to remove one a participant picked before it was discontinued).
    bloc_context = get_public_bloc_context(event, include_inactive=True)
    if not bloc_context:
        bloc_context = _empty_bloc_context()
    # One status/period rules set per period actually used among these
    # orders, so the popup's live preview can price each order against
    # its OWN period -- see registration_order_blocs_save's matching fix.
    period_rules_by_period_json = _period_rules_by_period_json(event, orders)

    context = {
        'event': event,
        'bloc_context': bloc_context,
        'period_rules_by_period_json': period_rules_by_period_json,
        'orders': orders,
        'status': status,
        'status_label': STATUS_LABELS.get(status, ''),
        'query': query,
        'submissions_count': len(orders),
        'total_revenue': total_revenue,
        'status_choices': STATUS_LABELS.items(),
        'status_breakdown': status_breakdown,
        'item_counts': item_counts,
        'bloc_item_options': bloc_item_options,
        'item_filters': item_filters,
        'workshop_item_ids': workshop_item_ids,
        'registrations_by_day': registrations_by_day,
        'show_events_link': show_events_link,
    }
    return render(request, 'dashboard/event_owner/submissions.html', context)


@never_cache
@login_required
def event_owner_new_registration(request, event_id):
    """
    Let the event owner create a brand-new registration directly, using
    the exact same paid-blocs catalog as the public form -- but with
    every status/item/workshop selectable even if it's since been
    deactivated (include_inactive=True), for cases like re-creating an
    old registration or a phone/in-person sign-up using a discontinued
    item. Unlike the public flow, this skips the email-verification-code
    step entirely (the owner is already authenticated) and creates the
    FormSubmission + RegistrationOrder immediately, mirroring
    views_blocs.finalize_paid_registration's own field set.
    """
    from django.core.files.storage import default_storage
    import uuid as uuid_lib
    from events.form_validation_service import create_participant_for_event, get_or_create_user_by_email
    from .models_form import FormSubmission

    event = get_object_or_404(Event, id=event_id)

    if not _can_access_event_orders(request.user, event):
        raise PermissionDenied("You do not have access to this event's submissions.")

    # Deliberately NOT filtered to is_active=True -- the owner can still
    # use this even if the public form has been closed.
    form_config = event.custom_forms.first()
    if not form_config:
        messages.error(request, "Aucun formulaire d'inscription n'est configuré pour cet événement.")
        return _redirect_to_submissions(request, event.id)

    bloc_context = get_public_bloc_context(event, include_inactive=True)
    if not bloc_context:
        messages.error(request, "Cet événement n'a aucune option d'inscription payante configurée.")
        return _redirect_to_submissions(request, event.id)

    errors = []

    if request.method == 'POST':
        # --- Informations fields -- same collection loop as the public
        # form (dashboard.views.public_form_view) ---
        form_data = {}
        email = None
        first_name = None
        last_name = None

        for field in form_config.fields_config:
            field_name = field.get('name')
            if field.get('type') == 'checkbox':
                value = request.POST.getlist(field_name)
            else:
                value = request.POST.get(field_name, '')

            if field.get('required') and (not value or (isinstance(value, list) and len(value) == 0)):
                errors.append(f'{field.get("label")} est requis.')
                continue

            form_data[field_name] = value
            if field_name == 'email' or field.get('type') == 'email':
                email = value.lower() if value else None
            elif field_name == 'first_name':
                first_name = value
            elif field_name == 'last_name':
                last_name = value

        if not email:
            errors.append("L'e-mail est requis.")
        if not first_name:
            errors.append('Le prénom est requis.')
        if not last_name:
            errors.append('Le nom est requis.')

        # --- Bloc/session selections -- same cardinality rules as
        # stage_paid_registration / registration_order_blocs_save ---
        selected_item_ids = []
        for bloc in bloc_context['custom_blocs']:
            picked = [v for v in request.POST.getlist(f"items_{bloc['key']}") if v]
            if bloc['select_mode'] == 'single' and len(picked) > 1:
                errors.append(f"Veuillez sélectionner une seule option dans {bloc['label']}.")
            if bloc['key'] == 'status' and len(picked) == 0:
                errors.append(f"Veuillez sélectionner une option dans {bloc['label']}.")
            selected_item_ids.extend(picked)

        selected_session_ids = [v for v in request.POST.getlist('sessions') if v]

        if not errors:
            result = compute_order(
                event=event, config=bloc_context['config'],
                selected_item_ids=selected_item_ids, selected_session_ids=selected_session_ids,
                on_date=timezone.now().date(), include_inactive=True, ignore_visibility_rules=True,
            )

            # Optional -- unlike the public flow, never required: the
            # owner is the one vouching for this registration directly.
            receipt_path = ''
            receipt = request.FILES.get('receipt_file')
            if receipt:
                receipt_path = default_storage.save(
                    f'registrations/receipts/{uuid_lib.uuid4().hex}_{receipt.name}', receipt
                )

            full_name = f"{first_name or ''} {last_name or ''}".strip()
            ip_address = (
                request.META.get('HTTP_X_FORWARDED_FOR', '').split(',')[0]
                if request.META.get('HTTP_X_FORWARDED_FOR') else request.META.get('REMOTE_ADDR')
            )
            user_agent = request.META.get('HTTP_USER_AGENT', '')[:1000]

            user = get_or_create_user_by_email(email, first_name, last_name)
            participant = create_participant_for_event(user, event)

            submission = FormSubmission.objects.create(
                form=form_config, data=form_data, email=email,
                ip_address=ip_address, user_agent=user_agent,
            )

            RegistrationOrder.objects.create(
                event=event, form_submission=submission, participant=participant,
                period_id=result['active_period_id'], full_name=full_name, email=email,
                items_snapshot=result['snapshot'], subtotals=result['subtotals'],
                distinct_blocs_count=result['distinct_blocs_count'],
                total_before_reduction=result['total_before_reduction'],
                period_discount_percent=result['period_discount_percent'],
                blocs_discount_percent=result['blocs_discount_percent'],
                total_discount_percent=result['total_discount_percent'],
                total_after_reduction=result['total_after_reduction'],
                receipt_file=receipt_path, ip_address=ip_address, user_agent=user_agent,
            )

            form_config.increment_submission_count()
            messages.success(request, f"Inscription créée pour {full_name or email}.")
            return _redirect_to_submissions(request, event.id)

    context = {
        'event': event,
        'form_config': form_config,
        'fields': form_config.fields_config,
        'bloc_context': bloc_context,
        'errors': errors,
    }
    return render(request, 'dashboard/event_owner/new_registration.html', context)


@never_cache
@login_required
def event_owner_export_excel(request, event_id):
    """
    Export the registrations table to Excel -- applies the exact same
    status/search/item filters as event_owner_submissions (duplicated
    rather than shared to avoid touching that view's already
    hard-won N+1/timeout-safe query shape), so the export always
    matches whatever's currently on screen.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    event = get_object_or_404(Event, id=event_id)

    if not _can_access_event_orders(request.user, event):
        raise PermissionDenied("You do not have access to this event's submissions.")

    status = request.GET.get('status', '').strip()
    query = request.GET.get('q', '').strip()
    item_filters = {
        'status': request.GET.get('item_status', '').strip(),
        'restauration': request.GET.get('item_restauration', '').strip(),
        'social_event': request.GET.get('item_social_event', '').strip(),
    }
    workshop_item_ids = [v for v in request.GET.getlist('item_workshops') if v]

    orders = list(
        RegistrationOrder.objects.filter(event=event).select_related(
            'period', 'form_submission__form',
        ).order_by('-created_at')
    )

    if status in STATUS_LABELS:
        orders = [o for o in orders if o.status == status]
    if query:
        ql = query.lower()
        orders = [o for o in orders if ql in (o.full_name or '').lower() or ql in (o.email or '').lower()]
    for bloc, item_id in item_filters.items():
        if item_id:
            orders = [o for o in orders if item_id in _bloc_item_ids(o, bloc)]
    if workshop_item_ids:
        wanted = set(workshop_item_ids)
        orders = [o for o in orders if _bloc_item_ids(o, 'workshops') & wanted]

    phone_field_name = None
    form_config = event.custom_forms.first()
    if form_config:
        phone_field_name = next(
            (f.get('name') for f in form_config.fields_config if f.get('type') == 'tel'), None,
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Inscriptions"

    header_fill = PatternFill(start_color="2D1B6B", end_color="2D1B6B", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    headers = [
        'Participant', 'E-mail', 'Téléphone', 'Statut choisi', 'Restauration',
        'Ateliers', 'Événement social', 'Prix (DZD)', "Statut de l'inscription",
        'Notes', 'Soumis le',
    ]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    for order in orders:
        bloc_names = _bloc_names(order)
        phone = (
            order.form_submission.data.get(phone_field_name, '')
            if order.form_submission and phone_field_name else ''
        )
        ws.append([
            order.full_name or '', order.email or '', phone,
            bloc_names['status'], bloc_names['restauration'], bloc_names['workshops'], bloc_names['social_event'],
            float(order.total_after_reduction), STATUS_LABELS.get(order.status, order.status),
            order.admin_notes or '', order.created_at.strftime('%Y-%m-%d %H:%M'),
        ])

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 22

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="inscriptions_{event.name}_{timezone.now().date()}.xlsx"'
    wb.save(response)
    return response


@never_cache
@login_required
@require_POST
def registration_status_save(request, order_id):
    """
    Move a registration between its 6 statuses. Confirmed triggers a real
    effect (see caisse.services -- creates a real paid CaisseTransaction +
    session access). Moving OFF Confirmed to ANY other status -- Cancelled
    or otherwise -- voids that transaction too (cancel_registration_order
    for Cancelled, unconfirm_registration_order for the rest), since none
    of them should leave a stale "paid" transaction behind for a
    registration that's no longer Confirmed. Between any two non-Confirmed
    statuses this is plain bookkeeping.
    """
    order = get_object_or_404(RegistrationOrder, id=order_id)
    if not _can_access_event_orders(request.user, order.event):
        raise PermissionDenied("You do not have access to this event's submissions.")

    new_status = request.POST.get('status', '').strip()
    if new_status not in STATUS_LABELS:
        messages.error(request, 'Statut invalide.')
        return _redirect_to_submissions(request, order.event_id)

    if new_status != order.status:
        if new_status == 'approved':
            try:
                confirm_registration_order(order, confirmed_by=request.user)
                messages.success(request, "Inscription confirmée -- le participant dispose désormais d'un accès payé réel.")
            except ValueError as exc:
                messages.error(request, str(exc))
        elif new_status == 'rejected':
            cancel_registration_order(order, cancelled_by=request.user)
            messages.success(request, 'Inscription annulée.')
        elif order.status == 'approved':
            # Moving off Confirmed to anything else isn't a plain label
            # change either -- void the real transaction it created,
            # same principle as Cancelled, just without marking the
            # registration itself as rejected.
            unconfirm_registration_order(order, new_status, changed_by=request.user)
            messages.success(
                request,
                "Statut de l'inscription mis à jour -- le paiement précédemment confirmé a été annulé.",
            )
        else:
            order.status = new_status
            order.reviewed_by = request.user
            order.reviewed_at = timezone.now()
            order.save(update_fields=['status', 'reviewed_by', 'reviewed_at'])
            messages.success(request, "Statut de l'inscription mis à jour.")

    return _redirect_to_submissions(request, order.event_id)


@never_cache
@login_required
@require_POST
def registration_notes_save(request, order_id):
    """Autosaved free-text note on a registration -- bookkeeping only,
    no side effects, so it responds with JSON instead of redirecting
    (avoids a full page reload/scroll-jump while someone is typing)."""
    order = get_object_or_404(RegistrationOrder, id=order_id)
    if not _can_access_event_orders(request.user, order.event):
        raise PermissionDenied("You do not have access to this event's submissions.")

    order.admin_notes = request.POST.get('admin_notes', '')
    order.save(update_fields=['admin_notes'])
    return JsonResponse({'ok': True})


@never_cache
@login_required
@require_POST
def registration_order_blocs_save(request, order_id):
    """
    Edit an existing registration's bloc selections (event owner "add more
    things" flow) -- recomputed through the exact same compute_order() the
    public registration form itself calls in finalize_paid_registration(),
    so the pricing/discount math is guaranteed identical, not a
    reimplementation that could quietly drift from it. Pinned to the
    order's own snapshotted period (period=order.period) rather than
    today's active period -- editing an old registration must keep
    pricing it at the rates that applied when it was placed, not
    whatever period happens to be running today.

    Also allowed once the order is 'approved', with real money/access
    effects: see caisse.services.resync_confirmed_registration_order,
    which cancels the CaisseTransaction confirm_registration_order
    created and replaces it with a new one for the updated total/items
    (never mutates a completed transaction's amount in place), and
    grants/revokes real SessionAccess for any paid workshop added/
    removed. Only possible when order.caisse_transaction is set -- i.e.
    it was confirmed through THIS dashboard's own confirm flow, not by a
    real physical caisse station, whose transaction can bundle several
    OTHER people's payments together and so isn't safe to touch here;
    those stay blocked, same restriction cancel_registration_order
    already applies. This is entirely separate from caisse's own
    on-the-day "add item" flow (caisse/views.py) which is untouched.

    Responds with JSON (called from the submissions page's blocs-editor
    modal via fetch) -- the caller reloads the page on success rather
    than this view re-rendering every affected cell itself.
    """
    order = get_object_or_404(RegistrationOrder, id=order_id)
    if not _can_access_event_orders(request.user, order.event):
        raise PermissionDenied("You do not have access to this event's submissions.")

    if order.status == 'approved' and not order.caisse_transaction_id:
        return JsonResponse({
            'ok': False,
            'error': "Cette inscription a été confirmée directement à la caisse (le paiement peut regrouper "
                     "plusieurs personnes) et ne peut pas être modifiée depuis ici.",
        }, status=400)

    bloc_context = get_public_bloc_context(order.event, include_inactive=True)
    if not bloc_context:
        return JsonResponse({
            'ok': False,
            'error': "Les options d'inscription de cet événement ne sont plus disponibles.",
        }, status=400)

    # Same cardinality rules as the public form's stage_paid_registration:
    # single-select blocs get at most one pick, Status is mandatory.
    selected_item_ids = []
    errors = []
    for bloc in bloc_context['custom_blocs']:
        picked = [v for v in request.POST.getlist(f"items_{bloc['key']}") if v]
        if bloc['select_mode'] == 'single' and len(picked) > 1:
            errors.append(f"Veuillez sélectionner une seule option dans {bloc['label']}.")
        if bloc['key'] == 'status' and len(picked) == 0:
            errors.append(f"Veuillez sélectionner une option dans {bloc['label']}.")
        selected_item_ids.extend(picked)

    selected_session_ids = [v for v in request.POST.getlist('sessions') if v]

    if errors:
        return JsonResponse({'ok': False, 'error': ' '.join(errors)}, status=400)

    # Pinned to the period this order was actually placed under (a
    # snapshot on the order itself -- see RegistrationOrder.period), not
    # whatever period happens to be active today: editing an old
    # registration must keep pricing it at its own period's rates.
    # ignore_visibility_rules=True: a status/period rule hiding an item
    # from a PARTICIPANT with that status must not also stop the owner
    # from manually granting/removing it here -- the popup's own JS
    # (submissions.html's applyPricingRules()) never hides a card for
    # this reason either, so what the owner can select here matches what
    # actually gets saved.
    result = compute_order(
        event=order.event, config=bloc_context['config'],
        selected_item_ids=selected_item_ids, selected_session_ids=selected_session_ids,
        on_date=timezone.now().date(), include_inactive=True, period=order.period,
        ignore_visibility_rules=True,
    )

    if order.status == 'approved':
        # Real transaction/access effects -- see resync_confirmed_registration_order,
        # which also saves order's pricing fields itself.
        resync_confirmed_registration_order(order, result, edited_by=request.user)
    else:
        order.period_id = result['active_period_id']
        order.items_snapshot = result['snapshot']
        order.subtotals = result['subtotals']
        order.distinct_blocs_count = result['distinct_blocs_count']
        order.total_before_reduction = result['total_before_reduction']
        order.period_discount_percent = result['period_discount_percent']
        order.blocs_discount_percent = result['blocs_discount_percent']
        order.total_discount_percent = result['total_discount_percent']
        order.total_after_reduction = result['total_after_reduction']
        order.save(update_fields=[
            'period', 'items_snapshot', 'subtotals', 'distinct_blocs_count',
            'total_before_reduction', 'period_discount_percent', 'blocs_discount_percent',
            'total_discount_percent', 'total_after_reduction',
        ])

    return JsonResponse({'ok': True})


@never_cache
@login_required
@require_POST
def send_payment_link_email(request, order_id):
    """
    Emails the participant a "préinscription" message. If the admin has
    configured a custom template, it's sent EXACTLY as authored (subject +
    body, only {{...}} variables substituted) -- nothing appended, nothing
    assumed. {{items}}/{{total}}/{{payment_link}} are available for the
    admin to place wherever they like, or omit entirely; it's their call.
    Only the built-in default (used until an admin bothers to customize)
    includes a sensible itemized breakdown + payment button automatically.
    Responds with JSON (called from the submissions modal via fetch, no
    page reload) rather than redirecting. Records payment_link_sent_at on
    success so the modal can keep showing "already sent" on any future
    visit, not just for as long as this one page stays open.
    """
    order = get_object_or_404(RegistrationOrder, id=order_id)
    if not _can_access_event_orders(request.user, order.event):
        raise PermissionDenied("You do not have access to this event's submissions.")

    event = order.event
    if not order.email:
        return JsonResponse({'ok': False, 'error': "Cette inscription n'a pas d'adresse e-mail."}, status=400)

    items_html = ''.join(
        f"<li>{it.get('name', '')} — {it.get('price', 0)} DZD</li>"
        for it in (order.items_snapshot or [])
    )
    if not items_html:
        items_html = '<li>Aucun article</li>'
    items_value = f"<ul>{items_html}</ul>"
    total_value = f"{order.total_after_reduction} DZD"

    template = get_event_email_template(event, 'payment_link')

    if template:
        subject = template.subject
        html_content = template.body_html or template.body
        replacements = {
            '{{participant_name}}': order.full_name or '',
            '{{event_name}}': event.name,
            '{{payment_link}}': event.payment_link or '',
            '{{items}}': items_value,
            '{{total}}': total_value,
        }
        for key, value in replacements.items():
            subject = subject.replace(key, value)
            html_content = html_content.replace(key, value)
    else:
        payment_link_html = ''
        if event.payment_link:
            payment_link_html = f"""
            <p>
                <a href="{event.payment_link}"
                   style="display:inline-block; padding:10px 20px; background:#2D1B6B; color:#ffffff; text-decoration:none; border-radius:6px;">
                   Procéder au paiement
                </a>
            </p>
            <p>Si le bouton ne fonctionne pas, copiez ce lien dans votre navigateur :<br>{event.payment_link}</p>
            """
        subject = f"Préinscription — {event.name}"
        html_content = f"""
        <html>
        <body style="font-family: Arial, sans-serif; color: #1a1a1a; line-height: 1.5;">
            <p>Bonjour {order.full_name or ''},</p>
            <p>Voici les informations de votre préinscription à <strong>{event.name}</strong>.</p>
            <p><strong>Articles sélectionnés :</strong></p>
            {items_value}
            <p><strong>Total à payer : {total_value}</strong></p>
            {payment_link_html}
            <p>Cordialement,<br>L'équipe {event.name}</p>
        </body>
        </html>
        """

    success, error, _ = send_email(
        to_email=order.email,
        subject=subject,
        html_content=html_content,
        to_name=order.full_name or None,
    )
    if not success:
        return JsonResponse({'ok': False, 'error': error or "Échec de l'envoi."}, status=502)

    order.payment_link_sent_at = timezone.now()
    order.save(update_fields=['payment_link_sent_at'])
    return JsonResponse({'ok': True, 'sent_at': order.payment_link_sent_at.isoformat()})


@never_cache
@login_required
@require_POST
def registration_delete(request, order_id):
    """
    Delete a registration and revoke the participant's access to THIS
    event only -- their global participant profile, and any access to
    other events they're separately registered for, is untouched.

    If it was already Confirmed (a real caisse transaction is attached),
    requires the extra confirm_paid=1 flag: a plain delete on a paid
    registration would silently leave real collected money with nothing
    on record explaining it.
    """
    order = get_object_or_404(RegistrationOrder, id=order_id)
    if not _can_access_event_orders(request.user, order.event):
        raise PermissionDenied("You do not have access to this event's submissions.")

    if order.status == 'approved' and not request.POST.get('confirm_paid'):
        messages.error(
            request,
            'Cette inscription était déjà confirmée (paiement réel enregistré). '
            'Annulez-la d\'abord, ou confirmez explicitement la suppression si vous comprenez cette action.',
        )
        return _redirect_to_submissions(request, order.event_id)

    event_id = order.event_id
    participant = order.participant

    if order.caisse_transaction_id and order.caisse_transaction.status == 'completed':
        cancel_label = request.user.get_full_name() or request.user.email
        order.caisse_transaction.cancel(cancelled_by=cancel_label, reason='Registration deleted by event owner.')

    if participant:
        # Only revoke this event's access if this was the participant's
        # LAST order for it -- someone with duplicate submissions (a real,
        # common case, see diagnose_registration_gap) still has another
        # valid order for this event after one duplicate is deleted, and
        # revoking their access anyway silently made them invisible to the
        # caisse even though their remaining order is still there.
        other_orders_remain = RegistrationOrder.objects.filter(
            event_id=event_id, participant=participant,
        ).exclude(id=order.id).exists()
        if not other_orders_remain:
            UserEventAssignment.objects.filter(
                user=participant.user, event_id=event_id, role='participant',
            ).update(is_active=False)
            ParticipantEventRegistration.objects.filter(participant=participant, event_id=event_id).delete()

    order.delete()
    messages.success(request, 'Inscription supprimée.')
    return _redirect_to_submissions(request, event_id)
